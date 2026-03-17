import io
import os
from datetime import datetime, date
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import xml.etree.ElementTree as ET

from database import get_db, Activity, Project

router = APIRouter(prefix="/api/projects", tags=["export"])

# Colour palette matching your programme
COLORS = {
    "summary_bg": "1F3864",      # Dark navy - summary rows
    "summary_fg": "FFFFFF",
    "critical_bg": "C00000",     # Red - critical activities
    "near_critical_bg": "E26B0A", # Orange - near critical
    "normal_bg": "2F75B6",        # Blue - normal activities
    "milestone_bg": "7030A0",     # Purple - milestones
    "header_bg": "1F3864",
    "header_fg": "FFFFFF",
    "complete_bar": "70AD47",     # Green progress
    "row_alt": "EBF3FB",          # Light blue alternating row
    "row_white": "FFFFFF",
    "border": "BFBFBF",
}


def parse_date(date_str: str):
    """Parse date string in various formats."""
    if not date_str:
        return None
    for fmt in ["%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d", "%a %d/%m/%y", "%a %d/%m/%Y"]:
        try:
            # Strip day prefix like "Mon " or "Fri "
            clean = date_str.strip()
            if len(clean) > 3 and clean[3] == " " and clean[:3].isalpha():
                clean = clean[4:]
            return datetime.strptime(clean, fmt).date()
        except ValueError:
            continue
    return None


@router.get("/{project_id}/export/excel")
def export_excel(project_id: int, db: Session = Depends(get_db)):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    activities = (
        db.query(Activity)
        .filter(Activity.project_id == project_id)
        .order_by(Activity.sort_order)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Construction Programme"

    # Page setup - A3 landscape
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 8  # A3
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # ── Header block ──────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = project.name.upper()
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=COLORS["header_fg"])
    title_cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:H2")
    sub_cell = ws["A2"]
    sub_cell.value = f"{project.address or ''} | {project.client or ''} | {project.revision or 'REV 1'}"
    sub_cell.font = Font(name="Calibri", size=10, color=COLORS["header_fg"])
    sub_cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16

    # ── Column headers ─────────────────────────────────────────────
    headers = ["ID", "Task Name", "Dur (d)", "Dur (w)", "Start", "Finish", "% Complete", "Resource Names"]
    col_widths = [6, 52, 8, 8, 13, 13, 12, 22]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(name="Calibri", bold=True, size=10, color=COLORS["header_fg"])
        cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            bottom=Side(style="medium", color=COLORS["border"]),
            right=Side(style="thin", color=COLORS["border"])
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[3].height = 18

    # ── Activity rows ──────────────────────────────────────────────
    thin = Side(style="thin", color=COLORS["border"])
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, act in enumerate(activities, start=4):
        ws.row_dimensions[row_idx].height = 16
        is_alt = (row_idx % 2 == 0)

        # Determine style
        if act.is_summary:
            bg = COLORS["summary_bg"]
            fg = COLORS["summary_fg"]
            bold = True
            size = 10
        elif act.is_milestone:
            bg = COLORS["milestone_bg"]
            fg = COLORS["summary_fg"]
            bold = True
            size = 9
        elif act.is_critical:
            bg = COLORS["critical_bg"]
            fg = COLORS["summary_fg"]
            bold = False
            size = 9
        elif act.is_near_critical:
            bg = COLORS["near_critical_bg"]
            fg = COLORS["summary_fg"]
            bold = False
            size = 9
        else:
            bg = COLORS["row_alt"] if is_alt else COLORS["row_white"]
            fg = "000000"
            bold = False
            size = 9

        fill = PatternFill("solid", fgColor=bg)

        # Indent task name
        indent = "  " * (act.indent_level or 0)
        name_val = f"{indent}{act.name}"

        row_data = [
            act.task_id,
            name_val,
            act.duration_days,
            act.duration_weeks,
            act.start_date,
            act.finish_date,
            f"{int(act.percent_complete or 0)}%" if not act.is_summary else "",
            act.resource_names or "",
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", bold=bold, size=size, color=fg)
            cell.fill = fill
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 7:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # ── Footer ─────────────────────────────────────────────────────
    footer_row = len(activities) + 5
    ws.merge_cells(f"A{footer_row}:H{footer_row}")
    footer_cell = ws[f"A{footer_row}"]
    footer_cell.value = (
        f"Generated by Construction Programme Manager | "
        f"Status Date: {project.status_date or 'N/A'} | "
        f"Printed: {datetime.now().strftime('%d/%m/%Y')} | "
        f"{project.working_days}-day working week"
    )
    footer_cell.font = Font(name="Calibri", size=8, italic=True, color="666666")
    footer_cell.alignment = Alignment(horizontal="left")

    # ── Freeze panes ───────────────────────────────────────────────
    ws.freeze_panes = "C4"

    # Stream response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{project.name} - {project.revision}.xlsx"
    filename = filename.replace("/", "-").replace("\\", "-")

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{project_id}/export/msproject")
def export_ms_project_xml(project_id: int, db: Session = Depends(get_db)):
    """Export as Microsoft Project XML format (.xml) - importable into MS Project."""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    activities = (
        db.query(Activity)
        .filter(Activity.project_id == project_id)
        .order_by(Activity.sort_order)
        .all()
    )

    root = ET.Element("Project")
    root.set("xmlns", "http://schemas.microsoft.com/project")

    # Project properties
    ET.SubElement(root, "Name").text = project.name
    ET.SubElement(root, "Title").text = project.name
    ET.SubElement(root, "Company").text = project.client or ""
    ET.SubElement(root, "CreationDate").text = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    ET.SubElement(root, "LastSaved").text = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    ET.SubElement(root, "ScheduleFromStart").text = "1"
    ET.SubElement(root, "DefaultDurationFormat").text = "7"  # days
    ET.SubElement(root, "DaysPerMonth").text = "20"
    ET.SubElement(root, "MinutesPerDay").text = str(project.working_days * 60)
    ET.SubElement(root, "MinutesPerWeek").text = str(project.working_days * 60 * 5)

    # Tasks
    tasks_el = ET.SubElement(root, "Tasks")

    for act in activities:
        task = ET.SubElement(tasks_el, "Task")
        ET.SubElement(task, "UID").text = str(act.id)
        ET.SubElement(task, "ID").text = str(act.task_id or act.id)
        ET.SubElement(task, "Name").text = act.name
        ET.SubElement(task, "OutlineLevel").text = str(act.indent_level or 0)
        ET.SubElement(task, "Summary").text = "1" if act.is_summary else "0"
        ET.SubElement(task, "Milestone").text = "1" if act.is_milestone else "0"
        ET.SubElement(task, "Critical").text = "1" if act.is_critical else "0"
        ET.SubElement(task, "PercentComplete").text = str(int(act.percent_complete or 0))

        dur_days = act.duration_days or 0
        ET.SubElement(task, "Duration").text = f"PT{int(dur_days * project.working_days * 60)}M"

        start_d = parse_date(act.start_date)
        finish_d = parse_date(act.finish_date)
        if start_d:
            ET.SubElement(task, "Start").text = start_d.strftime("%Y-%m-%dT08:00:00")
        if finish_d:
            ET.SubElement(task, "Finish").text = finish_d.strftime("%Y-%m-%dT17:00:00")

        if act.resource_names:
            ET.SubElement(task, "Notes").text = f"Resources: {act.resource_names}"

    xml_str = '<?xml version="1.0" encoding="UTF-8"?>\n' + ET.tostring(root, encoding="unicode")

    filename = f"{project.name} - {project.revision}.xml"
    filename = filename.replace("/", "-").replace("\\", "-")

    return StreamingResponse(
        io.StringIO(xml_str),
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
